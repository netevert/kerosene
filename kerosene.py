#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# About & license info ---------------------------------------------------------
"""
Version 1.0.1

ABOUT KEROSENE
==============
Kerosene is a lightweight, compact and intuitive application allowing users to
record and permanently store flight information in a personal microdatabase.
The program has a number of features:
    1 - A panel providing miscellaneous database and flight information
    2 - Navigable list allowing users to view/retrieve flights by date
    3 - A menu providing basic database management functionalities
    4 - A search by date tool to quickly retrieve flight Data
    5 - Facilities to export data to Excel and JSON formats
    6 - A dashboard to visually display flight routes on a map

LICENSE INFORMATION
===================
Kerosene version 1.0.1

The MIT License (MIT)

Copyright © 2017 ErrantBot

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.


ABOUT KEROSENE'S LICENSE
-------------------------
This license is what is generally known as the "MIT License",
aka "X11 License", "MIT/X Consortium License", "Expat License".
See http://opensource.org/licenses/MIT

This license is GPL-compatible.
See https://en.wikipedia.org/wiki/MIT_License
See http://www.gnu.org/licenses/license-list.html#GPLCompatibleLicenses

It is a permissive free software license, meaning that it permits reuse
within proprietary software provided all copies of the licensed software
include a copy of the MIT License terms and the copyright notice. Such
proprietary software retains its proprietary nature even though it
incorporates software under the MIT License.
Again see https://en.wikipedia.org/wiki/MIT_License

CREDITS
=======
As part of it's flight finder facility Kerosene uses tkentrycomplete.py
a tkinter widget that features autocompletion created by Mitja Martini
on 2008-11-29. This is a subclass of tkinter.Entry that features
autocompletion and the code can be found on Tkinter's Wiki page at
http://tkinter.unpythonic.net/wiki/AutocompleteEntry.

Kerosene is able to export data to .xlsx spreadsheet files thanks to openpyxl,
a Python library written by Eric Gazoni and also distributed under MIT license.
Openpyxl's website can be found at: https://openpyxl.readthedocs.org/en/latest/

In addition Kerosene's flame and Json icons are being used under a Creative
Commons Attribution 3.0 Unported License. More information on this license can
be found at https://creativecommons.org/licenses/by/3.0/
The authors of these icons can be found at http://www.aha-soft.com/ and
http://p.yusukekamiyamane.com/ respectively.

All other Icons used in the program, and their author are found at
http://www.famfamfam.com/
"""

# Imports ----------------------------------------------------------------------

# standard library modules
import time
import pickle
from collections import Counter
import datetime as dt
from idlelib.ToolTip import ToolTip
from itertools import chain
import json
import os
from shutil import copy
import sqlite3
from threading import Thread
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog as fd
import tkinter.messagebox as box
import webbrowser as web
import sys
if getattr(sys, 'frozen', False):
    os.environ['BASEMAPDATA'] = os.path.join(os.path.dirname(sys.executable),
                                             'data')

# third-party modules
import matplotlib
from matplotlib.figure import Figure
from mpl_toolkits.basemap import Basemap
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, \
    NavigationToolbar2TkAgg
import numpy as np
from openpyxl import Workbook

# Constants --------------------------------------------------------------------

YEARS = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025,
         2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037,
         2038, 2039, 2040, 2041, 2042, 2043, 2044, 2045, 2046, 2047, 2048, 2049,
         2050, 2051, 2052, 2053, 2054, 2055, 2056, 2057, 2058, 2059, 2060, 2061,
         2062, 2063, 2064, 2065, 2066, 2067, 2068, 2069, 2070, 2071, 2072, 2073,
         2074, 2075, 2076, 2077, 2078, 2079, 2080, 2081, 2082, 2083, 2084, 2085,
         2086, 2087, 2088, 2089, 2090, 2091, 2092, 2093, 2094, 2095, 2096, 2097,
         2098, 2099, 2100, 2101, 2102, 2103, 2104, 2105, 2106, 2107, 2108, 2109,
         2110, 2111, 2112, 2113, 2114, 2115, 2116, 2117, 2118, 2119, 2120, 2121,
         2122, 2123, 2124, 2125, 2126, 2127, 2128, 2129, 2130, 2131, 2132, 2133,
         2134, 2135, 2136, 2137, 2138, 2139, 2140, 2141, 2142, 2143, 2144, 2145,
         2146, 2147, 2148, 2149, 2150, 2151, 2152, 2153, 2154, 2155, 2156, 2157,
         2158, 2159, 2160, 2161, 2162, 2163, 2164, 2165, 2166, 2167, 2168, 2169,
         2170, 2171, 2172, 2173, 2174, 2175, 2176, 2177, 2178, 2179, 2180, 2181,
         2182, 2183, 2184, 2185, 2186, 2187, 2188, 2189, 2190, 2191, 2192, 2193,
         2194, 2195, 2196, 2197, 2198, 2199, 2200, 2201, 2202, 2203, 2204, 2205,
         2206, 2207, 2208, 2209, 2210, 2211, 2212, 2213, 2214, 2215, 2216, 2217,
         2218, 2219, 2220, 2221, 2222, 2223, 2224, 2225, 2226, 2227, 2228, 2229,
         2230, 2231, 2232, 2233, 2234, 2235, 2236, 2237, 2238, 2239, 2240, 2241,
         2242, 2243, 2244, 2245, 2246, 2247, 2248, 2249, 2250]

MONTHS = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11",
          "12"]

DAYS = ["01", "02", "03", "04", "05", "06",
        "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18",
        "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30",
        "31"]

HOURS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
         21, 22, 23, 24]

MINS_SECS = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18,
             19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35,
             36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52,
             53, 54, 55, 56, 57, 58, 59, 60]

NOW = dt.datetime.now()
ICON_LINK_1 = 'http://www.aha-soft.com/'
ICON_LINK_2 = 'http://www.famfamfam.com/'
ICON_LINK_3 = 'http://p.yusukekamiyamane.com/'
CREDITS = (
    'This program is being distributed under the terms of the MIT License'
    ' a copy of which has been included within the distribution folder'
    ' you have downloaded from the program\'s website. The license can be'
    ' found in the file named LICENSE.txt. Should this file not be '
    'present within the distribution folder you are kindly invited to '
    'consult the terms of the MIT License at'
    ' opensource.org/licenses/MIT.')

IATA_CODE_LIST = []


# Search box autocompletion code -----------------------------------------------


# noinspection PyClassHasNoInit
class AutocompleteEntry(tk.Entry):
    """ tkentrycomplete.py
        A tkinter widget that features autocompletion.
        Created by Mitja Martini on 2008-11-29.
        Subclass of tkinter.Entry that features autocompletion.

        To enable autocompletion use set_completion_list(list) to define
        a list of possible strings to hit.
        To cycle through hits use down and up arrow keys.
        """

    def set_completion_list(self, completion_list):
        self._completion_list = completion_list
        self._hits = []
        self._hit_index = 0
        self.position = 0
        self.bind('<KeyRelease>', self.handle_keyrelease)

    def autocomplete(self, delta=0):
        """ autocomplete the Entry,
        delta may be 0/1/-1 to cycle through possible hits
        :param delta: """
        if delta:  # need to delete selection otherwise we would /
            self.delete(self.position, tk.END)  # fix the current position
        else:  # set position to end so selection starts where textentry ended
            self.position = len(self.get())
        # collect hits
        _hits = []
        for element in self._completion_list:
            if element.startswith(self.get().lower()):
                _hits.append(element)
        # if we have a new hit list, keep this in mind
        if _hits != self._hits:
            self._hit_index = 0
            self._hits = _hits
        # only allow cycling if we are in a known hit list
        if _hits == self._hits and self._hits:
            self._hit_index = (self._hit_index + delta) % len(self._hits)
        # now finally perform the auto completion
        if self._hits:
            self.delete(0, tk.END)
            self.insert(0, self._hits[self._hit_index])
            self.select_range(self.position, tk.END)

    def handle_keyrelease(self, event):
        """ event handler for the keyrelease event on this widget
        :param event:
        """
        tkinter_umlauts = ['odiaeresis', 'adiaeresis', 'udiaeresis',
                           'Odiaeresis', 'Adiaeresis', 'Udiaeresis', 'ssharp']
        if event.keysym == "BackSpace":
            self.delete(self.index(tk.INSERT), tk.END)
            self.position = self.index(tk.END)
        if event.keysym == "Left":
            if self.position < self.index(tk.END):  # delete the selection
                self.delete(self.position, tk.END)
            else:
                self.position -= 1  # delete one character
                self.delete(self.position, tk.END)
        if event.keysym == "Right":
            self.position = self.index(tk.END)  # go to end (no selection)
        if event.keysym == "Down":
            self.autocomplete(1)  # cycle to next hit
        if event.keysym == "Up":
            self.autocomplete(-1)  # cycle to previous hit
        # perform normal autocomplete if event is a single key or an umlaut
        if len(event.keysym) == 1 or event.keysym in tkinter_umlauts:
            self.autocomplete()


# Matplotlib custom toolbar ----------------------------------------------------

# noinspection PyAbstractClass
class CustomToolbar(NavigationToolbar2TkAgg):
    """Customized Matplotlib toolbar with famfam icons added."""

    def __init__(self, canvas_, parent_):
        self.toolitems = (
            ('Home', 'Reset original view', 'house', 'home'),
            ('Back', 'Back to previous view', 'arrow_left', 'back'),
            ('Forward', 'Forward to next view', 'arrow_right', 'forward'),
            (None, None, None, None),
            ('Pan', 'Pan axes with left mouse, zoom with right', 'arrow_inout',
             'pan'),
            ('Zoom', 'Zoom to rectangle', 'zoom_in', 'zoom'),
            (None, None, None, None),
            ('Save', 'Save map image', 'disk', 'save_figure'),
        )
        NavigationToolbar2TkAgg.__init__(self, canvas_, parent_)


# Statistics tab class ---------------------------------------------------------


class StatisticsTab(tk.Frame):
    """Creates and builds database statistics tab.

       This class is responsible for the building and operation of the
       program's statistics tab, displaying useful database information to
       the user."""

    def __init__(self, master=None):
        """Initializes the tab's tk.Frame object, sets an image constant and
           launches tab build command."""
        super().__init__()
        self.exitimage = tk.PhotoImage(file='data/icons/tab_delete.png')
        self.build()

    def build(self):
        """Initializes and builds all of the tab's widgets."""
        self.label_frame_stats = tk.LabelFrame(self)
        self.label_frame_stats.pack(side='top', expand='yes', fill='both',
                                    padx=1, pady=1)

        # create and grid top destination field
        self.top_destination_label = tk.Label(self.label_frame_stats,
                                              text="Most frequented airport:")
        self.top_destination_label.grid(row=0, column=0, sticky="w")
        self.entry_top_destination = tk.Entry(self.label_frame_stats, width=34,
                                              state=tk.DISABLED,
                                              disabledforeground="black")
        self.entry_top_destination.grid(row=0, column=1,
                                        columnspan=4, sticky="w")

        # create and grid an exit tab button
        self.button_exit = ttk.Button(self.label_frame_stats,
                                      image=self.exitimage,
                                      command=self.close_tab)
        self.button_exit.grid(row=0, column=4, sticky="e")
        self.button_tip = ToolTip(self.button_exit,
                                  ["Exit", "statistics", "tab"])

        # create and grid total flight hours field
        self.flight_hours_label = tk.Label(self.label_frame_stats,
                                           text="Flight hours accumulated:")
        self.flight_hours_label.grid(row=1, column=0, sticky="w")
        self.entry_flight_hours = tk.Entry(self.label_frame_stats, width=15,
                                           state=tk.DISABLED,
                                           disabledforeground="black")
        self.entry_flight_hours.grid(row=1, column=1, sticky="w")

        # create and grid average flight time field
        self.average_time_label = tk.Label(self.label_frame_stats,
                                           text="Average flight time:")
        self.average_time_label.grid(row=1, column=3, sticky="w")
        self.entry_average_time = tk.Entry(self.label_frame_stats, width=15,
                                           state=tk.DISABLED,
                                           disabledforeground="black")
        self.entry_average_time.grid(row=1, column=4, sticky="w")

        # create and grid total flights logged in database field
        self.total_logs_label = tk.Label(self.label_frame_stats,
                                         text="Flights logged:")
        self.total_logs_label.grid(row=2, column=3, sticky="w")
        self.entry_total_logs = tk.Entry(self.label_frame_stats, width=15,
                                         state=tk.DISABLED,
                                         disabledforeground="black")
        self.entry_total_logs.grid(row=2, column=4, sticky="w")

        # create and grid total number of airports visited field
        self.label_airports_visited = tk.Label(self.label_frame_stats,
                                               text="Number of airports visited:")
        self.label_airports_visited.grid(row=2, column=0, sticky="w")
        self.entry_airports_visited = tk.Entry(self.label_frame_stats, width=15,
                                               state=tk.DISABLED,
                                               disabledforeground="black")
        self.entry_airports_visited.grid(row=2, column=1, sticky="w")

        # create and grid top carrier field
        self.label_top_carrier = tk.Label(self.label_frame_stats,
                                          text="Top carrier:")
        self.label_top_carrier.grid(row=3, column=3, sticky="w")
        self.entry_top_carrier = tk.Entry(self.label_frame_stats, width=15,
                                          state=tk.DISABLED,
                                          disabledforeground="black")
        self.entry_top_carrier.grid(row=3, column=4, sticky="w")

        # create and gridlongest flight field
        self.label_longest_flight = tk.Label(self.label_frame_stats,
                                             text="Longest flight duration:")
        self.label_longest_flight.grid(row=3, column=0, sticky="w")
        self.entry_longest_flight = tk.Entry(self.label_frame_stats, width=15,
                                             state=tk.DISABLED,
                                             disabledforeground="black")
        self.entry_longest_flight.grid(row=3, column=1, sticky="w")

        # create and grid most travelled year field
        self.label_year_most_travelled = tk.Label(self.label_frame_stats,
                                                  text="Year with most flights:")
        self.label_year_most_travelled.grid(row=4, column=0, sticky="w")
        self.entry_year_most_travelled = tk.Entry(self.label_frame_stats,
                                                  width=15,
                                                  state=tk.DISABLED,
                                                  disabledforeground="black")
        self.entry_year_most_travelled.grid(row=4, column=1, sticky="w")

        # create and grid top aircraft boarded field
        self.label_top_aircraft = tk.Label(self.label_frame_stats,
                                           text="Top aircraft:")
        self.label_top_aircraft.grid(row=4, column=3, sticky="w")
        self.entry_top_aircraft = tk.Entry(self.label_frame_stats, width=15,
                                           state=tk.DISABLED,
                                           disabledforeground="black")
        self.entry_top_aircraft.grid(row=4, column=4, sticky="w")

        # create and grid total number of cities visited field
        self.label_cities_visited = tk.Label(self.label_frame_stats,
                                             text="Cities visited:")
        self.label_cities_visited.grid(row=5, column=0, sticky="w")
        self.entry_cities_visited = tk.Entry(self.label_frame_stats,
                                             width=15,
                                             state=tk.DISABLED,
                                             disabledforeground="black")
        self.entry_cities_visited.grid(row=5, column=1, sticky="w")

        # create and grid city most travelled to field
        self.label_top_city = tk.Label(self.label_frame_stats,
                                       text="Top city:")
        self.label_top_city.grid(row=5, column=3, sticky="w")
        self.entry_top_city = tk.Entry(self.label_frame_stats, width=15,
                                       state=tk.DISABLED,
                                       disabledforeground="black")
        self.entry_top_city.grid(row=5, column=4, sticky="w")

        # give padding to all widgets
        for child in self.label_frame_stats.winfo_children():
            child.grid_configure(padx=1, pady=1)

        # refresh tab statistical information
        self.update_data()

    def update_data(self):
        """ Uploads/refreshes new Data on current flight Data labelframe """
        # set up database
        database = sqlite3.connect('data/database.kr')
        cursor = database.cursor()

        # enable statistics tab entry fields
        self.entry_total_logs.config(state="normal")
        self.entry_flight_hours.config(state="normal")
        self.entry_average_time.config(state="normal")
        self.entry_top_destination.config(state="normal")
        self.entry_airports_visited.config(state="normal")
        self.entry_top_carrier.config(state="normal")
        self.entry_longest_flight.config(state="normal")
        self.entry_year_most_travelled.config(state="normal")
        self.entry_top_aircraft.config(state="normal")
        self.entry_cities_visited.config(state="normal")
        self.entry_top_city.config(state="normal")

        # empty statistics tab entry fields
        self.entry_total_logs.delete(0, tk.END)
        self.entry_flight_hours.delete(0, tk.END)
        self.entry_average_time.delete(0, tk.END)
        self.entry_top_destination.delete(0, tk.END)
        self.entry_airports_visited.delete(0, tk.END)
        self.entry_top_carrier.delete(0, tk.END)
        self.entry_longest_flight.delete(0, tk.END)
        self.entry_year_most_travelled.delete(0, tk.END)
        self.entry_top_aircraft.delete(0, tk.END)
        self.entry_cities_visited.delete(0, tk.END)
        self.entry_top_city.delete(0, tk.END)

        # reset database panel variables
        no_hours = dt.timedelta(hours=0, minutes=0, seconds=0)  # dummy time
        self.total_flights_logged = 0
        self.number_of_airports = 0
        self.no_of_cities = 0
        self.average_flight_time = no_hours
        self.total_flight_hours = no_hours
        self.longest_flight = no_hours
        self.favourite_airport = ""
        self.year_most_flown = ""
        self.top_carrier = ""
        self.top_plane = ""
        self.top_city = ""

        # Data extraction and re-calculation algorithms ------------------------
        try:
            flight_hours = dt.timedelta(hours=0, minutes=0, seconds=0)
            longest = dt.timedelta(hours=0, minutes=0, seconds=0)
            airports_visited = []
            carrier_list = []
            planes_list = []
            cities_list = []
            years_list = []
            form = "%H:%M:%S"
            for flight in cursor.execute('''SELECT * FROM flight_data'''):

                # convert flight_duration database field to timedelta object
                # and add to total flight hours timedelta object
                flight_duration = self.string_to_timedelta(flight[5])
                flight_hours += flight_duration

                date = flight[0][:4]  # strip off day and month, keep only year
                years_list.append(date)
                carrier_list.append(flight[8])
                planes_list.append(flight[2])
                city = flight[13]
                city_2 = flight[14]
                self.total_flights_logged += 1
                if flight_duration > longest:
                    longest = flight_duration
                if city not in cities_list:  # avoids double counting cities
                    cities_list.append(city)
                    self.no_of_cities += 1
                if city_2 not in cities_list:  # avoids double counting cities
                    cities_list.append(city_2)
                    self.no_of_cities += 1

            # chained list comprehension including both departure airport and 
            # destination airport lists
            deps_des = chain([flight[6] for flight in
                              cursor.execute('''SELECT * FROM flight_data''')],
                             [flight[7] for flight in
                              cursor.execute('''SELECT * FROM flight_data''')])

            # chained list comprehension containing both departure city and
            # destination city lists
            cities = chain([flight[13] for flight in
                            cursor.execute('''SELECT * FROM flight_data''')],
                           [flight[14] for flight in
                            cursor.execute('''SELECT * FROM flight_data''')])

            # find most common data
            count = Counter(deps_des)
            top_of_count = count.most_common(1)
            self.favourite_airport = str(top_of_count[0][0])
            count_carriers = Counter(carrier_list)
            top_carrier = count_carriers.most_common(1)
            self.top_carrier = str(top_carrier[0][0])
            count_cities = Counter(cities)
            top_city = count_cities.most_common(1)
            self.top_city = str(top_city[0][0])
            count_plane_models = Counter(planes_list)
            top_plane = count_plane_models.most_common(1)
            self.top_plane = str(top_plane[0][0])
            count_years = Counter(years_list)
            top_year = count_years.most_common(1)
            self.year_most_flown = str(top_year[0][0])

            # calculate average flight time
            self.average_flight_time = str(flight_hours //
                                           self.total_flights_logged)
            flight_hours = self.format_timedelta(flight_hours)
            self.total_flight_hours = str(flight_hours)
            self.number_of_airports = str(self.total_flights_logged * 2)
            self.longest_flight = str(longest)

            # free up memory
            del flight_hours
            del airports_visited[:]
            del carrier_list[:]
            del planes_list[:]
            del cities_list[:]
            del years_list[:]

        except IndexError:
            pass

        finally:
            # Update widgets
            self.entry_total_logs.insert(0, str(self.total_flights_logged))
            self.entry_flight_hours.insert(0, str(self.total_flight_hours))
            self.entry_average_time.insert(0, str(self.average_flight_time))
            self.entry_top_destination.insert(0, self.favourite_airport)
            self.entry_airports_visited.insert(0, self.number_of_airports)
            self.entry_top_carrier.insert(0, self.top_carrier)
            self.entry_longest_flight.insert(0, self.longest_flight)
            self.entry_year_most_travelled.insert(0, self.year_most_flown)
            self.entry_top_aircraft.insert(0, self.top_plane)
            self.entry_cities_visited.insert(0, self.no_of_cities)
            self.entry_top_city.insert(0, self.top_city)

            # disable statistics tab entry fields
            self.entry_total_logs.config(state="disabled")
            self.entry_flight_hours.config(state="disabled")
            self.entry_average_time.config(state="disabled")
            self.entry_top_destination.config(state="disabled")
            self.entry_airports_visited.config(state="disabled")
            self.entry_top_carrier.config(state="disabled")
            self.entry_longest_flight.config(state="disabled")
            self.entry_year_most_travelled.config(state="disabled")
            self.entry_top_aircraft.config(state="disabled")
            self.entry_cities_visited.config(state="disabled")
            self.entry_top_city.config(state="disabled")

            # close database and update status
            database.close()
            kerosene.label_status['text'] = "Idle..."

    def close_tab(self):
        """Destroys statistics tab and re-enables the statistics menu option."""
        self.destroy()
        kerosene.enable_menu_statistics()

    @staticmethod
    def string_to_timedelta(string=''):
        """This function converts datetime string representations saved in
        the database."""
        split_string = string.split(':')
        converted_string = dt.timedelta(hours=int(split_string[0]),
                                        minutes=int(split_string[1]),
                                        seconds=int(split_string[2]))
        return converted_string

    @staticmethod
    def format_timedelta(time_to_format):
        """Formats timedelta objects to display time in a
        HOURS:MINUTES:SECONDS format."""
        minutes, seconds = divmod(time_to_format.seconds +
                                  time_to_format.days * 86400, 60)
        hours, minutes = divmod(minutes, 60)
        return '{:d}:{:02d}:{:02d}'.format(hours, minutes, seconds)


# Routemap tab class -----------------------------------------------------------


class RoutemapTab(tk.Frame):
    """Creates and manages Matplotlib plotting tab.

       This class creates and manages the program's tab responsible for plotting
       flight routes or airports visited on a Matplotlib map and displaying the
       results to the user.

       WARNING: The routes plotted on Matplotlib's maps are not the actual
                navigation routes followed by the plane during flight. They are
                simply a great circle visual representation connecting the
                flight's departure and destination airports."""

    def __init__(self, master=None, key=None, is_route=True, is_pickled=None):
        """Initializes the tab's underlying tk.Frame object and sets constants.

           This initialization function carries out an important task besides
           invoking super() and setting constants. Based on the user's menu
           selection recorded in the is_route argument it pulls the appropriate
           data from the database for either plotting individual flight route
           data or displaying all visited airports on the map.

           Args:
               key: The flight record's database key, used to retrieve the
                    flight's departure and destination airport latitude and
                    longitude.
               is_route: If True the initialization method pulls the flight
                         flight record's latitude and longitude data. If False
                          __init__ will extract latitude/longitude for all
                          airports visited."""
        super().__init__()
        self.worldimage = tk.PhotoImage(file='data/icons/world_delete.png')
        self.key = key
        self.is_route = is_route
        self.is_pickled = is_pickled
        self.flight_number = ""

        if self.is_route:  # pull latitude and longitude of a single flight record

            # connect to database
            database = sqlite3.connect('data/database.kr')
            try:
                cursor = database.cursor()
                fetch_row = cursor.execute('''SELECT * FROM flight_data
                                              WHERE date=?''', (self.key,))
                record = fetch_row.fetchone()

                # fetch departure and destination latitude and longitude
                self.from_latitude = float(record[11])
                self.from_longitude = float(record[12])
                self.to_latitude = float(record[15])
                self.to_longitude = float(record[16])

                # fetch flight number
                self.flight_number = record[1]

            except sqlite3.Error:
                box.showwarning('Error', 'Oops, something went wrong!'
                                         '\nThe database appears not'
                                         '\nto be working properly.')
            finally:

                # close database
                database.close()

            # invoke method that plots the single flight route

            start = time.time()
            self.build_route()

        else:  # pull latitude and longitude of all visited airports

            database = sqlite3.connect('data/database.kr')
            try:
                cursor = database.cursor()

                # chained list comprehension including both departure latitude
                # and destination latitude
                self.lats = chain([float(flight[11]) for flight in
                                   cursor.execute(
                                       '''SELECT * FROM flight_data''')],
                                  [float(flight[15]) for flight in
                                   cursor.execute(
                                       '''SELECT * FROM flight_data''')])

                # chained list comprehension including both departure longitude
                # and destination longitude
                self.lons = chain([float(flight[12]) for flight in
                                   cursor.execute(
                                       '''SELECT * FROM flight_data''')],
                                  [float(flight[16]) for flight in
                                   cursor.execute(
                                       '''SELECT * FROM flight_data''')])

                # set latitude and longitude constants
                self.lats = tuple(self.lats)
                self.lons = tuple(self.lons)

            except sqlite3.Error:
                box.showwarning('Error', 'Oops, something went wrong!'
                                         '\nThe database appears not'
                                         '\nto be working properly.')
            finally:
                database.close()

            # invoke method that plots all airports on the map
            self.build_airports_map()

    @staticmethod
    def build_map(is_pickled=None):
        """Function to load pickled map; if one doesn't exist it's created."""
        if not is_pickled:
            matplotlib.use('TkAgg')
            fig = Figure()

            # set map background color
            fig.patch.set_facecolor('blue')

            ax1 = fig.add_subplot(111)
            m = Basemap(projection='robin',
                        lat_0=43.0000, lon_0=12.0000,
                        resolution='c', area_thresh=100000,
                        ax=ax1)
            # build map and adjust settings
            m.drawcoastlines(linewidth=1, color="yellow")
            m.drawcountries(color="blue")
            m.fillcontinents(color='blue', lake_color="blue")
            m.drawmapboundary(linewidth=2, color="yellow", fill_color="blue")
            m.drawmeridians(np.arange(0, 360, 30), color="yellow")
            m.drawparallels(np.arange(-90, 90, 30), color="yellow")
            # self.map.bluemarble()

            map_pickle = pickle.dumps((m, fig))
            return m, fig
        else:
            return pickle.loads(is_pickled)

    def build_route(self):
        """Initializes map and plots individual flight route data"""

        # initialize map
        m, fig = self.build_map(self.is_pickled)

        # set departure and destination latitude/longitude using __init__
        # constants
        fromlat = self.from_latitude
        fromlon = self.from_longitude
        tolat = self.to_latitude
        tolon = self.to_longitude

        # mark departure and destination airports
        lats = [fromlat, tolat]
        lons = [fromlon, tolon]
        x, y = m(lons, lats)
        m.plot(x, y, 'rv', markersize=8)

        # draw creat curcle between departure and destination
        m.drawgreatcircle(fromlon, fromlat, tolon, tolat, linewidth=2,
                          color='r')

        # create exit button
        self.button_quit = ttk.Button(self, image=self.worldimage,
                                      command=self.close_tab)
        self.button_quit.pack(side="top", fill="y", padx=2, anchor="e")
        self.button_tip = ToolTip(self.button_quit, ["Exit", "route", "tab"])

        # create canvas
        self.canvas = FigureCanvasTkAgg(fig, master=self)
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1,
                                         padx=2, pady=2)

        # initialize and update custom toolbar
        toolbar = CustomToolbar(self.canvas, self)
        toolbar.update()

        # show canvas
        self.canvas.show()

        # noinspection PyProtectedMember
        self.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True,
                                   padx=2, pady=2)
        # update status label
        kerosene.label_status["text"] = "Viewing route plot for flight " \
                                        "{}".format(self.flight_number)

    def build_airports_map(self):
        # todo: convert like build map
        """Initializes map and plots airport positions."""
        # initialize map
        m, fig = self.build_map(self.is_pickled)

        # set airports departure and destination latitude/longitude using
        # __init__ constants
        lats = self.lats
        lons = self.lons
        x, y = m(lons, lats)

        # plot data
        m.plot(x, y, 'rv', markersize=8)

        # create exit button
        self.button_quit = ttk.Button(self, image=self.worldimage,
                                      command=self.close_tab)
        self.button_quit.pack(side="top", fill="y", padx=2, anchor="e")
        self.button_tip = ToolTip(self.button_quit, ["Exit", "route", "tab"])

        # create canvas
        self.canvas = FigureCanvasTkAgg(fig, master=self)
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1,
                                         padx=2, pady=2)

        # initialize and update custom toolbar
        toolbar = CustomToolbar(self.canvas, self)
        toolbar.update()

        # show canvas
        self.canvas.show()

        # noinspection PyProtectedMember
        self.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True,
                                   padx=2, pady=2)

        # update status label
        kerosene.label_status["text"] = "Viewing all airports visited on map"

    def close_tab(self):
        """Quits routemap tab."""
        self.destroy()
        kerosene.enable_menu_options()
        try:
            if kerosene.tab_stats.winfo_exists():

                # ensure view statistics menu option remains disabled
                kerosene.disable_menu_statistics()
            else:
                pass

        # ignore attribute error when statistics tab is not open
        except AttributeError:
            pass
        finally:
            # update status label
            kerosene.label_status["text"] = "Idle..."


# Uploader and editor tab class ------------------------------------------------


class UploaderEditorTab(tk.Frame):
    """Creates and manages upload and edit data tab.

       This is the main class responsible for interacting with the
       database. It allows the user to upload new flight data or
       re-edit currently stored data."""

    def __init__(self, master=None, is_uploader=False,
                 key=None, *args, **kwargs):
        """Crates frame for uploader and editor tabs and sets constants.

           Args:
               is_uploader: If True the tab will be initialized to upload data.
                            If False it will be initialized to edit the existing
                            flight record selected in the main tab.
               key: The flight record's database key, used to retrieve the
                    necessary flight record data."""
        super().__init__()

        # set constants
        self.uploadimage = tk.PhotoImage(file='data/icons/database_add.png')
        self.editimage = tk.PhotoImage(file='data/icons/database_edit.png')
        self.is_uploader = is_uploader
        self.YEAR = dt.datetime.now().year
        self.DAY = dt.datetime.now().day
        self.MONTH = [month for month in MONTHS
                      if str(dt.datetime.now().month) in month]
        self.key = key

        if not self.is_uploader:

            # extract the record's key and sanitize data to prepare for
            # insertion in edit tab fields
            database = sqlite3.connect('data/database.kr')
            cursor = database.cursor()
            fetch_row = cursor.execute('''SELECT * FROM flight_data
                                              WHERE date=?''', (self.key,))
            record = fetch_row.fetchone()

            # if database key is duplicate, extract only the first 10 characters
            # to discard "-(VALUE)" tag appended to key. Keys are saved in a
            # YYYY-MM-DD format. If two flights are recorded in the same date a
            # "-(VALUE)" tag will be added to the key, e.g. 2015-05-4-(2).
            # This duplicate tag has to be stripped away before splitting the
            # key in year, month and day variables; otherwise the variables
            # cannot be used to update related tab combobox widgets and will
            # cause errors.
            if len(record[0]) > 10:
                _time = record[0][:10]  # cut off the duplicate flag

                # in duplicate keys where the DD value is a single digit,
                # e.g. 2015-05-4-(2), cutting off the flag will result in the
                # string: 2015-05-4-. When splitting and unpacking the key the
                # remaining '-' also needs to be removed to avoid causing errors
                if _time[9] == '-':
                    _time = time[:9]

                # split and unpack key value
                self.yr, self.mnth, self.dy = _time.split('-')
            else:

                # split and unpack normally
                self.yr, self.mnth, self.dy = record[0].split('-')

            # set constants
            self.flight_number = record[1]
            self.plane_model = record[2]
            self.take_off_hour, self.take_off_minutes, self.take_off_seconds = (
                record[3].split(':'))
            self.landing_hour, self.landing_minutes, self.landing_seconds = (
                record[4].split(':'))
            self.departure = record[6]
            self.destination = record[7]
            self.carrier = record[8]
            self.iata = record[9]
            self.iata_2 = record[10]
            self.latitude = record[11]
            self.longitude = record[12]
            self.city = record[13]
            self.city_2 = record[14]
            database.close()

        self.build()

    def build(self):
        """"Initializes and grids UploaderEditorTab widgets.

            Builds frame widgets and, based on the contents of is_uploader,
            implements uploading/editing logic."""
        # ------------------- create flight date form --------------------------

        # create labelframe container
        self.frame_date = tk.LabelFrame(self)
        self.frame_date.grid(row=0, column=0, padx=1, pady=1, sticky="w")

        # flight year fields
        self.label_year = tk.Label(self.frame_date, text='Date: ')
        self.label_year.grid(row=0, column=0, padx=1, pady=1, sticky='w')
        self.var_year = tk.StringVar()
        self.year = ttk.Combobox(self.frame_date, textvariable=self.var_year,
                                 values=YEARS, state='readonly', width=5)
        self.year.grid(row=0, column=1, padx=1, pady=1)
        self.tip_year = ToolTip(self.year, ["Select", "a", "year"])

        # use current year if tab is opened to upload new data, otherwise use
        # the year extracted from the record's key
        if self.is_uploader:
            self.year.set(self.YEAR)
        else:
            self.year.set(self.yr)

        # separate year field from month field
        self.label_dash = tk.Label(self.frame_date,
                                   text=" - ").grid(row=0, column=2)

        # flight month fields
        self.var_month = tk.StringVar()
        self.month = ttk.Combobox(self.frame_date, textvariable=self.var_month,
                                  value=MONTHS, state='readonly', width=2)
        self.month.grid(row=0, column=3)
        self.tip_month = ToolTip(self.month, ["Select", "a", "month"])

        # use current month if tab is opened to upload new data, otherwise use
        # the month value extracted from the record's key
        if self.is_uploader:
            self.month.set(self.MONTH[0])
        else:
            self.month.set(self.mnth)

        # separate month field from day field
        self.label_dash_2 = tk.Label(self.frame_date,
                                     text=" - ").grid(row=0, column=4)

        # flight day field
        self.var_day = tk.StringVar()
        self.day = ttk.Combobox(self.frame_date, textvariable=self.var_day,
                                value=DAYS, state='readonly', width=2)
        self.day.grid(row=0, column=5, padx=1)
        self.tip_day = ToolTip(self.day, ["Select", "a", "day"])

        # use current day if tab is opened to upload new data, otherwise use
        # the day value extracted from the record's key
        if self.is_uploader:
            self.day.set(self.DAY)
        else:
            self.day.set(self.dy)

        # create a tk.Label with an empty string buffer to stretch
        # self.frame_date to the left side of self.button_main: this is a very
        # inelegant and, most probably, incorrect solution and will be changed
        # if an alternative is found.
        buffer = " " * 75
        self.lbl = tk.Label(self.frame_date, text=buffer)
        self.lbl.grid(row=0, column=6, columnspan=2)

        # create a button to upload/ edit data or, alternatively, exit
        self.button_main = ttk.Button(self, command=self.process_data,
                                      image=None)
        self.button_main.grid(row=0, column=1, sticky="e")

        # set main button text and tooltip according to one of the tab's
        # opening modes (upload/edit)
        if self.is_uploader:
            self.button_main['image'] = self.uploadimage
            self.button_tip = ToolTip(self.button_main, ["Upload", "data", "or",
                                                         "exit"])
        else:
            self.button_main["image"] = self.editimage
            self.button_tip = ToolTip(self.button_main, ["Edit", "data", "or",
                                                         "exit"])

        # ----------- create departure/destination information form ------------

        # create departure information labelframe container
        self.frame_airports = tk.LabelFrame(self)
        self.frame_airports.grid(row=1, column=0, padx=1, pady=1, columnspan=2)

        self.label_departure = tk.Label(self.frame_airports,
                                        text="Take off:")
        self.label_departure.grid(row=1, column=0, columnspan=2, sticky="W")

        # flight departure time: hour field
        self.var_hour = tk.IntVar()
        self.hours = ttk.Combobox(self.frame_airports,
                                  textvariable=self.var_hour,
                                  values=HOURS, state='readonly', width=4)
        self.hours.grid(row=1, column=1, sticky='e')
        self.hours_tip = ToolTip(self.hours, ["Select", "take", "off", "hour"])
        self.label_dash_3 = tk.Label(self.frame_airports,
                                     text=" - ").grid(row=1, column=2)

        # flight departure time: minutes field
        self.var_mins = tk.IntVar()
        self.minutes = ttk.Combobox(self.frame_airports,
                                    textvariable=self.var_mins, value=MINS_SECS,
                                    state='readonly', width=24)
        self.minutes.grid(row=1, column=3)
        self.tip_mins = ToolTip(self.minutes, ["Select", "minutes"])
        self.label_dash_4 = tk.Label(self.frame_airports,
                                     text=" - ").grid(row=1, column=4)

        # flight departure time: seconds field
        self.var_sec = tk.IntVar()
        self.seconds = ttk.Combobox(self.frame_airports,
                                    textvariable=self.var_sec,
                                    value=MINS_SECS, state='readonly', width=15)
        self.seconds.grid(row=1, column=5, padx=1, pady=1)
        self.sec_tip = ToolTip(self.seconds, ["Select", "seconds"])

        # flight departure information: IATA code field
        self.label_iata = tk.Label(self.frame_airports, text="Iata code")
        self.label_iata.grid(row=2, column=0)
        self.entry_iata = ttk.Combobox(self.frame_airports,
                                       value=IATA_CODE_LIST,
                                       width=4)
        self.entry_iata.grid(row=2, column=1)

        # flight departure information: airport name field
        self.label_airport = tk.Label(self.frame_airports,
                                      text="Airport")
        self.label_airport.grid(row=2, column=2)
        self.entry_airport = tk.Entry(self.frame_airports, width=27)
        self.entry_airport.grid(row=2, column=3)

        # flight departure information: airport city field
        self.label_city = tk.Label(self.frame_airports,
                                   text="City")
        self.label_city.grid(row=2, column=4)
        self.entry_city = tk.Entry(self.frame_airports, width=19)
        self.entry_city.grid(row=2, column=5)

        # create destination information labelframe container
        self.label_destination = tk.Label(self.frame_airports,
                                          text="Landing:")
        self.label_destination.grid(row=3, column=0, columnspan=2, sticky="W")

        # flight destination time: hour field
        self.var_hour_2 = tk.IntVar()
        self.hours_2 = ttk.Combobox(self.frame_airports,
                                    textvariable=self.var_hour_2, values=HOURS,
                                    state='readonly', width=4)
        self.hours_2.grid(row=3, column=1, sticky='e')
        self.hours_tip_2 = ToolTip(self.hours_2,
                                   ["Select", "landing", "hour"])
        self.label_dash_4 = tk.Label(self.frame_airports,
                                     text=" - ").grid(row=3, column=2)

        # flight destination time: minutes field
        self.var_mins_2 = tk.IntVar()
        self.minutes_2 = ttk.Combobox(self.frame_airports,
                                      textvariable=self.var_mins_2,
                                      value=MINS_SECS, state='readonly',
                                      width=24)
        self.minutes_2.grid(row=3, column=3, sticky="w")
        self.tip_mins_2 = ToolTip(self.minutes_2, ["Select", "minutes"])
        self.label_dash_5 = tk.Label(self.frame_airports,
                                     text=" - ").grid(row=3, column=4)

        # flight destination time: seconds field
        self.var_sec_2 = tk.IntVar()
        self.seconds_2 = ttk.Combobox(self.frame_airports,
                                      textvariable=self.var_sec_2,
                                      value=MINS_SECS, state='readonly',
                                      width=15)
        self.seconds_2.grid(row=3, column=5, padx=1, pady=1)
        self.tip_sec_2 = ToolTip(self.seconds_2, ["Select", "seconds"])

        # flight destination information: IATA code field
        self.label_iata_2 = tk.Label(self.frame_airports, text="Iata code")
        self.label_iata_2.grid(row=4, column=0, padx=1, pady=1)
        self.entry_iata_2 = ttk.Combobox(self.frame_airports,
                                         value=IATA_CODE_LIST,
                                         width=4)
        self.entry_iata_2.grid(row=4, column=1)

        # flight destination information: airport name field
        self.label_airport_2 = tk.Label(self.frame_airports,
                                        text="Airport")
        self.label_airport_2.grid(row=4, column=2)
        self.entry_airport_2 = tk.Entry(self.frame_airports, width=27)
        self.entry_airport_2.grid(row=4, column=3)

        # flight destination information: airport city field
        self.label_city_2 = tk.Label(self.frame_airports,
                                     text="City")
        self.label_city_2.grid(row=4, column=4)
        self.entry_city_2 = tk.Entry(self.frame_airports, width=19)
        self.entry_city_2.grid(row=4, column=5)

        # ------------------ create flight information form --------------------

        # create flight information labelframe container
        self.frame_plane = tk.LabelFrame(self)
        self.frame_plane.grid(row=3, column=0, sticky="we", columnspan=2)

        # flight information: flight number field
        self.label_flight = tk.Label(self.frame_plane, text="Flight Number:")
        self.label_flight.grid(row=0, column=0, padx=1, pady=1)
        self.entry_flight = tk.Entry(self.frame_plane, width=9)
        self.entry_flight.grid(row=0, column=1)

        # flight information field: aircraft type field
        self.label_aircraft = tk.Label(self.frame_plane, text="Aircraft:")
        self.label_aircraft.grid(row=0, column=2)
        self.entry_aircraft = tk.Entry(self.frame_plane, width=15)
        self.entry_aircraft.grid(row=0, column=3)

        # flight information field: carrier field
        self.label_carrier = tk.Label(self.frame_plane, text="Carrier:")
        self.label_carrier.grid(row=0, column=4)
        self.entry_carrier = tk.Entry(self.frame_plane, width=20)
        self.entry_carrier.grid(row=0, column=5)

        # if the tab is opened in editor mode update flight information fields
        # to show the database records for the flight being edited
        if not self.is_uploader:
            # insert data into departure time fields
            self.hours["state"] = "default"
            self.minutes["state"] = "default"
            self.seconds["state"] = "default"
            self.hours.delete(0, tk.END)
            self.minutes.delete(0, tk.END)
            self.seconds.delete(0, tk.END)
            self.hours.insert(0, self.take_off_hour)
            self.minutes.insert(0, self.take_off_minutes)
            self.seconds.insert(0, self.take_off_seconds)
            self.hours["state"] = "readonly"
            self.minutes["state"] = "readonly"
            self.seconds["state"] = "readonly"

            # insert data into departure airport fields
            self.entry_iata.insert(0, self.iata)
            self.entry_airport.insert(0, self.departure)
            self.entry_city.insert(0, self.city)

            # insert data into landing time fields
            self.hours_2["state"] = "default"
            self.minutes_2["state"] = "default"
            self.seconds_2["state"] = "default"
            self.hours_2.delete(0, tk.END)
            self.minutes_2.delete(0, tk.END)
            self.seconds_2.delete(0, tk.END)
            self.hours_2.insert(0, self.landing_hour)
            self.minutes_2.insert(0, self.landing_minutes)
            self.seconds_2.insert(0, self.landing_seconds)
            self.hours_2["state"] = "readonly"
            self.minutes_2["state"] = "readonly"
            self.seconds_2["state"] = "readonly"

            # insert data into destination airport fields
            self.entry_iata_2.insert(0, self.iata_2)
            self.entry_airport_2.insert(0, self.destination)
            self.entry_city_2.insert(0, self.city_2)

            # insert data into flight information fields
            self.entry_flight.insert(0, self.flight_number)
            self.entry_aircraft.insert(0, self.plane_model)
            self.entry_carrier.insert(0, self.carrier)

        # add padding
        for child in self.winfo_children():
            child.grid_configure(padx=1, pady=1)

        # bind iata comboboxes to methods that automatically retrieve airport
        # and city associated with the selected IATA code
        self.entry_iata.bind('<<ComboboxSelected>>', self.populate_iata_1)
        self.entry_iata_2.bind('<<ComboboxSelected>>', self.populate_iata_2)

    def process_data(self, event=None):
        """Function controlling upload of new or edited data.

           This function controls whether the main button in the
           UploaderEditor tab is to implement the database_upload
           or upload_changes command. The first is associated with
           opening the tab in 'upload' mode; the second is used
           when the tab is opened in 'editor' mode. The object's
           is_uploader paramenter is used to determine in what mode
           the tab has been launched (using boolean values)."""
        if self.is_uploader:
            self.database_upload()
        else:
            self.upload_changes()

    def database_upload(self):
        """Uploads new Data to database, asks user to confirm
            Data input before uploading"""
        database = sqlite3.connect('data/database.kr')
        try:
            cursor = database.cursor()

            # check if information has been inputted
            if not self.var_year.get() \
                    or not self.var_month.get() \
                    or not self.var_day.get() \
                    or not self.var_hour.get() \
                    or not self.var_hour_2.get() \
                    or not self.entry_iata.get() \
                    or not self.entry_airport.get() \
                    or not self.entry_city.get() \
                    or not self.entry_iata_2.get() \
                    or not self.entry_airport_2.get() \
                    or not self.entry_city_2.get() \
                    or not self.entry_flight.get() \
                    or not self.entry_aircraft.get() \
                    or not self.entry_carrier.get():
                query_1 = box.askyesno(title='Warning',
                                       message="All fields must be completed!"
                                               "\nDo you want to stop uploading"
                                               "\nnew flight data?")
                if query_1 is True:
                    self.close_tab()  # exit uploader tab
                else:
                    pass  # return to uploader tab
            else:

                # create database key string
                yr = str(self.year.get())
                mon = self.month.get()
                d = self.day.get()
                date = "%s-%s-%s" % (yr, mon, d)

                # check for database key duplicates and, if found, generate a
                # duplicate key with format YY-MM-DD-(VALUE)
                dates = []
                for saved_date in cursor.execute('SELECT date FROM '
                                                 'flight_data ORDER BY date'):
                    dates.append(saved_date[0])

                if date in dates:
                    count = [i for i in dates if date in i]
                    date += "-(%d)" % len(count)

                # get user inputs
                name = self.entry_aircraft.get()
                number = self.entry_flight.get()
                take_off = dt.timedelta(hours=self.var_hour.get(),
                                        minutes=self.var_mins.get(),
                                        seconds=self.var_sec.get())
                landing = dt.timedelta(hours=self.var_hour_2.get(),
                                       minutes=self.var_mins_2.get(),
                                       seconds=self.var_sec_2.get())
                ftime = landing - take_off  # calculate flight time
                start = self.entry_airport.get()
                iata = self.entry_iata.get()
                city = self.entry_city.get()
                end = self.entry_airport_2.get()
                iata_2 = self.entry_iata_2.get()
                city_2 = self.entry_city_2.get()
                carr = self.entry_carrier.get()

                # fetch latitude and longitude for departure and destination
                latitude_dep = self.fetch("latitude", iata)
                longitude_dep = self.fetch("longitude", iata)
                latitude_des = self.fetch("latitude", iata_2)
                longitude_des = self.fetch("longitude", iata_2)
                text = ("\n"
                        "    Are you sure you want\n"
                        "    to upload this Data?\n"
                        "\n"
                        "    Date: %s\n"
                        "    Aircraft name: %s\n"
                        "    Flight number: %s\n"
                        "    Flight time: %s\n"
                        "    Departure: %s\n"
                        "    Destination: %s\n"
                        "    Carrier: %s\n"
                        "                    "
                        ) % (date, name, number, ftime, start, end,
                             carr)
                query_2 = box.askyesno(title='Confirm', message=text)
                if query_2 is True:

                    # upload data
                    box.showinfo(title='Upload Successfull',
                                 message='Flight %s'
                                         ' successfully '
                                         'added!' % number)
                    cursor.execute("""INSERT INTO flight_data (
                                            date,
                                            flight_number,
                                            plane_model,
                                            take_off_time,
                                            landing_time,
                                            flight_duration,
                                            departure,
                                            destination,
                                            carrier,
                                            iata_dep,
                                            iata_des,
                                            latitude_dep,
                                            longitude_dep,
                                            city_dep,
                                            city_des,
                                            latitude_des,
                                            longitude_des)
                                            VALUES(?,?,?,?,?,?,?,?,
                                                   ?,?,?,?,?,?,?,?,?)""",
                                   (str(date),
                                    str(number),
                                    str(name),
                                    str(take_off),
                                    str(landing),
                                    str(ftime),
                                    str(start),
                                    str(end),
                                    str(carr),
                                    str(iata),
                                    str(iata_2),
                                    str(latitude_dep),
                                    str(longitude_dep),
                                    str(city),
                                    str(city_2),
                                    str(latitude_des),
                                    str(longitude_des)))
                    database.commit()

                    # exit uploader tab
                    self.close_tab()
                else:
                    query_3 = box.askyesno('Termination',
                                           'Do you want to terminate '
                                           'data upload?')
                    if query_3 is True:
                        self.close_tab()  # exit uploader tab
                    else:
                        pass
        except sqlite3.Error:
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
        finally:
            database.close()

    def upload_changes(self):
        """Uploads new Data to database, asks user to confirm
            Data input before uploading"""
        database = sqlite3.connect('data/database.kr')
        try:
            cursor = database.cursor()

            # check if information has been inputted
            if not self.var_year.get() \
                    or not self.var_month.get() \
                    or not self.var_day.get() \
                    or not self.var_hour.get() \
                    or not self.var_hour_2.get() \
                    or not self.entry_iata.get() \
                    or not self.entry_airport.get() \
                    or not self.entry_city.get() \
                    or not self.entry_iata_2.get() \
                    or not self.entry_airport_2.get() \
                    or not self.entry_city_2.get() \
                    or not self.entry_flight.get() \
                    or not self.entry_aircraft.get() \
                    or not self.entry_carrier.get():

                query_1 = box.askyesno(title='Warning',
                                       message="All fields must be completed!"
                                               "\nDo you want to stop making"
                                               "\nchanges to fight data?")
                if query_1 is True:
                    self.close_tab()  # exit editor tab
                else:
                    pass  # return to editor tab
            else:

                # create key string
                yr = str(self.year.get())
                mon = self.month.get()
                d = self.day.get()
                date = "%s-%s-%s" % (yr, mon, d)

                # check for database key duplicates and, if found, generate a
                # duplicate key with format YY-MM-DD-(VALUE)
                dates = []
                for saved_date in cursor.execute('SELECT date FROM '
                                                 'flight_data ORDER BY date'):
                    dates.append(saved_date[0])

                if date in dates:
                    count = [i for i in dates if date in i]
                    date += "-(%d)" % len(count)

                # get user input
                name = self.entry_aircraft.get()
                number = self.entry_flight.get()
                take_off = dt.timedelta(hours=self.var_hour.get(),
                                        minutes=self.var_mins.get(),
                                        seconds=self.var_sec.get())
                landing = dt.timedelta(hours=self.var_hour_2.get(),
                                       minutes=self.var_mins_2.get(),
                                       seconds=self.var_sec_2.get())
                ftime = landing - take_off  # calculate flight times
                start = self.entry_airport.get()
                iata = self.entry_iata.get()
                city = self.entry_city.get()
                end = self.entry_airport_2.get()
                iata_2 = self.entry_iata_2.get()
                city_2 = self.entry_city_2.get()
                carr = self.entry_carrier.get()

                # fetch latitude and longitude for departure and destination
                latitude_dep = self.fetch("latitude", iata)
                longitude_dep = self.fetch("longitude", iata)
                latitude_des = self.fetch("latitude", iata_2)
                longitude_des = self.fetch("longitude", iata_2)
                text = ("\n"
                        "    Are you sure you want\n"
                        "    to upload this Data?\n"
                        "\n"
                        "    Date: %s\n"
                        "    Aircraft name: %s\n"
                        "    Flight number: %s\n"
                        "    Flight time: %s\n"
                        "    Departure: %s\n"
                        "    Destination: %s\n"
                        "    Carrier: %s\n"
                        "                    "
                        ) % (date, name, number, ftime, start, end,
                             carr)
                query_2 = box.askyesno(title='Confirm', message=text)
                if query_2 is True:

                    # upload edited data
                    box.showinfo(title='Upload Successfull',
                                 message='Flight %s '
                                         'successfully '
                                         'updated!' % number)
                    cursor.execute("""UPDATE flight_data SET
                                            date = ?,
                                            flight_number = ?,
                                            plane_model  = ?,
                                            take_off_time = ?,
                                            landing_time = ?,
                                            flight_duration = ?,
                                            departure = ?,
                                            destination = ?,
                                            carrier = ?,
                                            iata_dep = ?,
                                            iata_des = ?,
                                            latitude_dep = ?,
                                            longitude_dep = ?,
                                            city_dep = ?,
                                            city_des = ?,
                                            latitude_des = ?,
                                            longitude_des = ?
                                            WHERE date = ?""", (str(date),
                                                                str(number),
                                                                str(name),
                                                                str(take_off),
                                                                str(landing),
                                                                str(ftime),
                                                                str(start),
                                                                str(end),
                                                                str(carr),
                                                                str(iata),
                                                                str(iata_2),
                                                                str(
                                                                    latitude_dep),
                                                                str(
                                                                    longitude_dep),
                                                                str(city),
                                                                str(city_2),
                                                                str(
                                                                    latitude_des),
                                                                str(
                                                                    longitude_des),
                                                                self.key))
                    database.commit()

                    # exit editor tab
                    self.close_tab()
                else:
                    query_3 = box.askyesno('Termination',
                                           'Do you want to terminate '
                                           'data editing?')
                    if query_3 is True:
                        self.close_tab()  # exit editor tab
                    else:
                        pass
        except sqlite3.Error as e:
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
        finally:
            database.close()

    def populate_iata_1(self, event=None):
        """Fetches IATA codes from aiports database and populates comboboxes.

           This function is used to auto-populate the airport name
           and airport city entries in the upload tab when the user
           makes a selection in the iata code combobox. The function
           accesses the internal airports database and retrieves the
           name and city records using the iata combobox value as
           reference."""
        database = sqlite3.connect('data/airports_data.sqlite')
        iata = self.entry_iata.get()
        self.entry_airport.delete(0, tk.END)
        self.entry_city.delete(0, tk.END)
        try:
            command = "SELECT name, city FROM airports WHERE iata=?"
            cursor = database.cursor()
            fetch_row = cursor.execute(command, (iata,))
            record = fetch_row.fetchone()
            name, city = record[0], record[1]
            self.entry_airport.insert(0, name)
            self.entry_city.insert(0, city)
        except sqlite3.Error:
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
            pass
        finally:
            database.close()

    def populate_iata_2(self, event=None):
        """Fetches IATA codes from aiports database and populates comboboxes.

            The function is identical to populate_iata_1, like that
            function this is also used to auto-populate the airport name
            and airport city entries in the upload tab when the user
            makes a selection in the iata code combobox. The function
            accesses the internal airports database and retrieves the
            name and city records using the iata combobox value as
            reference """
        database = sqlite3.connect('data/airports_data.sqlite')
        iata = self.entry_iata_2.get()
        self.entry_airport_2.delete(0, tk.END)
        self.entry_city_2.delete(0, tk.END)
        try:
            command = "SELECT name, city FROM airports WHERE iata=?"
            cursor = database.cursor()
            fetch_row = cursor.execute(command, (iata,))
            record = fetch_row.fetchone()
            name, city = record[0], record[1]
            self.entry_airport_2.insert(0, name)
            self.entry_city_2.insert(0, city)
        except sqlite3.Error:
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
            pass
        finally:
            database.close()

    def close_tab(self):
        """Destroys the currently opened tab and resets GUI"""
        self.destroy()
        kerosene.populate_list()
        kerosene.switch_to_tab_1()
        kerosene.enable_menu_options()
        try:
            if kerosene.tab_stats.winfo_exists():

                # update stats if tab exists
                kerosene.tab_stats.update_data()

                # ensure view statistics menu option remains disabled
                kerosene.disable_menu_statistics()
            else:
                pass

        # ignore attribute error when statistics tab is not open
        except AttributeError:
            pass

    @staticmethod
    def fetch(data_type="", iata_arg=None):
        """ Fetches latitude or longitude from airports sqlite database.

            Args:
                data_type: The data_type argument tells the functions whether
                           latitude or longitude data is to be returned.
                iata_arg: takes the iata code supplied and uses it to select the
                          appropriate data cell from the approriate row in the
                          airports database."""
        database = sqlite3.connect('data/airports_data.sqlite')
        if data_type == "latitude":
            command = "SELECT latitude FROM airports WHERE iata =?"
        else:
            command = "SELECT longitude FROM airports WHERE iata =?"
        try:
            cursor = database.cursor()
            fetch_row = cursor.execute(command, (iata_arg,))
            record = fetch_row.fetchone()
            latitude = record[0]
            return latitude
        except sqlite3.Error:
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
            pass
        finally:
            database.close()


# Credits tool class -----------------------------------------------------------


class CreditsTool(tk.Toplevel):
    """Opens a new window providing information regarding author, program
    version, license and links to used Icons."""

    def __init__(self, master=None, *args, **kwargs):
        """Initializes Toplevel object and builds credit interface."""
        super().__init__()
        self.build()

    def build(self):
        """Initializes and builds application widgets."""
        self.img_1 = tk.PhotoImage(file='data/icons/python-powered-h-50x65.png')
        self.img_2 = tk.PhotoImage(file='data/icons/sqlite.png')

        # create main credits label
        self.lbl_info = tk.Label(self, text='Kerosene'
                                            '\ncopyright © {year}'
                                            '\nErrantBot'
                                            '\nversion '
                                            '1.0.1'.format(year=NOW.year),
                                 font=('courier', 10, 'normal'))

        self.lbl_info.grid(row=0, column=0, sticky='w', padx=1, pady=1)

        # create python logo credit
        self.lbl_logo = tk.Label(self, image=self.img_1, cursor='hand2')
        self.lbl_logo.grid(row=0, column=1, sticky='e', padx=1, pady=1)
        self.logo_tip = ToolTip(self.lbl_logo,
                                ["Check", "out", "www.python.org"])

        # create SQLite logo credit
        self.sql_logo = tk.Label(self, image=self.img_2, cursor='hand2')
        self.sql_logo.grid(row=0, column=2, sticky='e', padx=1, pady=1)
        self.sql_tip = ToolTip(self.sql_logo,
                               ["Check", "out", "www.sqlite.org"])

        # create credits text labelframe
        self.credits_labelframe = tk.LabelFrame(self,
                                                text='License information',
                                                foreground='brown')
        self.credits_labelframe.grid(row=1, column=0, padx=1, pady=1,
                                     sticky='we', columnspan=3)

        # create credits text display
        self.credits_display = tk.Text(self.credits_labelframe)
        self.scrollbar = tk.Scrollbar(self.credits_labelframe)
        self.credits_display.grid(row=0, column=0, padx=1, pady=1, columnspan=1)
        self.credits_display.insert(0.0, CREDITS)
        self.credits_display.config(state=tk.DISABLED, wrap=tk.WORD,
                                    height=5, width=33,
                                    font=('courier', 8, 'normal'))
        self.scrollbar.grid(row=0, column=1, padx=1, pady=1, sticky='ens')
        self.credits_display.config(yscrollcommand=self.scrollbar.set)
        self.scrollbar.config(command=self.credits_display.yview)

        # create hyperlink labelframe
        self.credits_labelframe_2 = tk.LabelFrame(self, text='Icon credits',
                                                  foreground='brown')
        self.credits_labelframe_2.grid(row=2, column=0, padx=1, pady=1,
                                       sticky='we', columnspan=3)

        # create hyperlink labels and grid them
        self.lbl_link_1 = tk.Label(self.credits_labelframe_2,
                                   text='Flame icon: ' + ICON_LINK_1,
                                   cursor='hand2')
        self.lbl_link_1.grid(row=0, column=0, padx=0, pady=0, sticky='w')

        self.lbl_link_2 = tk.Label(self.credits_labelframe_2,
                                   text='Other icons: ' + ICON_LINK_2,
                                   cursor='hand2')
        self.lbl_link_2.grid(row=1, column=0, padx=0, pady=0, sticky='w')

        self.lbl_link_3 = tk.Label(self.credits_labelframe_2,
                                   text='Json icon: ' + ICON_LINK_3,
                                   cursor='hand2')
        self.lbl_link_3.grid(row=2, column=0, padx=0, pady=0, sticky='w')

        # bind link labels to hyperlink functions
        self.lbl_link_1.bind('<Button-1>', self.hyperlink_1)
        self.lbl_link_2.bind('<Button-1>', self.hyperlink_2)
        self.lbl_link_3.bind('<Button-1>', self.hyperlink_3)
        self.lbl_logo.bind('<Button-1>', self.hyperlink_4)
        self.sql_logo.bind('<Button-1>', self.hyperlink_5)

    @staticmethod
    def hyperlink_1(event=None):
        """Opens link to specified URL for credit purpose."""
        web.open_new(r"http://www.aha-soft.com/")

    @staticmethod
    def hyperlink_2(event=None):
        """Opens link to specified URL for credit purposes."""
        web.open_new(r"http://www.famfamfam.com/")

    @staticmethod
    def hyperlink_3(event=None):
        """Opens link to specified URL for credit purposes."""
        web.open_new(r"http://p.yusukekamiyamane.com/")

    @staticmethod
    def hyperlink_4(event=None):
        """Opens link to specified URL for credit purposes."""
        web.open_new(r"https://www.python.org/")

    @staticmethod
    def hyperlink_5(event=None):
        """Opens link to specified URL for credit purposes."""
        web.open_new(r"https://www.sqlite.org/")


# Main program Gui -------------------------------------------------------------


class Gui(tk.Frame):
    """Builds Kerosene's GUI."""

    def __init__(self, master=None, *args, **kwargs):
        """Initializes main gui tk.Frame object.

           Initializes Frame object, some class variables, the program menubar
           and program widgets. Additionally if a database file does not exist one
           is created."""
        super().__init__()

        # Initialize program images
        self.uploadimage = tk.PhotoImage(file='data/icons/database_add.png')
        self.deleteimage = tk.PhotoImage(file='data/icons/database_delete.png')
        self.editimage = tk.PhotoImage(file='data/icons/database_edit.png')
        self.exportimage = tk.PhotoImage(file='data/icons/folder_database.png')
        self.importimage = tk.PhotoImage(file='data/icons/database_go.png')
        self.detailimage = tk.PhotoImage(file='data/icons/view_detail.png')
        self.statsimage = tk.PhotoImage(file='data/icons/chart_bar.png')
        self.worldimage = tk.PhotoImage(file='data/icons/world_go.png')
        self.mapimage = tk.PhotoImage(file='data/icons/map_go.png')
        self.findimage = tk.PhotoImage(file='data/icons/magnifier.png')
        self.clearimage = tk.PhotoImage(file='data/icons/textfield_delete.png')
        self.generateimage = tk.PhotoImage(file='data/icons/database_table.png')
        self.jsonimage = tk.PhotoImage(file='data/icons/json.png')
        self.helpimage = tk.PhotoImage(file='data/icons/help.png')
        self.aboutimage = tk.PhotoImage(file='data/icons/information.png')

        # Check if database file exhists, if not create one
        if os.path.exists('data/database.kr'):
            pass
        else:
            database = sqlite3.connect('data/database.kr')
            try:
                cursor = database.cursor()
                cursor.execute("""CREATE TABLE flight_data (date TEXT,
                                                        flight_number TEXT,
                                                        plane_model TEXT,
                                                        take_off_time TEXT,
                                                        landing_time TEXT,
                                                        flight_duration TEXT,
                                                        departure TEXT,
                                                        destination TEXT,
                                                        carrier TEXT,
                                                        iata_dep TEXT,
                                                        iata_des TEXT,
                                                        latitude_dep TEXT,
                                                        longitude_dep TEXT,
                                                        city_dep TEXT,
                                                        city_des TEXT,
                                                        latitude_des TEXT,
                                                        longitude_des TEXT)""")
                database.commit()
            except sqlite3.Error:
                box.showwarning('Error', 'Oops, something went wrong!'
                                         '\nThe database appears not'
                                         '\nto be working properly.')
            finally:
                database.close()

        # build menu, main interface widgets and cache base map
        self.build_menu()
        self.build_interface()
        self.map_pickle = self.pickle_map()

    def build_menu(self):
        """Initializes and builds program menubar."""
        self.top = tk.Menu(self)

        # create file menu
        # commands: export database, import database, exit
        self.file = tk.Menu(self.top, tearoff=False)
        self.file.add_command(label='Export Database', accelerator='Ctrl+E',
                              command=self.export_database,
                              image=self.exportimage, compound=tk.LEFT,
                              underline=0)
        self.file.add_command(label='Import Database', accelerator='Ctrl+I',
                              command=self.import_database,
                              image=self.importimage, compound=tk.LEFT,
                              underline=0)
        self.file.add_separator()
        self.file.add_command(label='Exit', command=self.quit_program,
                              underline=0)
        self.top.add_cascade(label='File', menu=self.file, underline=0)

        # create edit menu
        # commands: upload flight, edit flight, delete flight
        self.edit = tk.Menu(self.top, tearoff=False)
        self.edit.add_command(label='Upload Flight',
                              image=self.uploadimage, compound=tk.LEFT,
                              command=self.upload_flight, underline=0)
        self.edit.add_command(label='Edit Flight',
                              image=self.editimage, compound=tk.LEFT,
                              command=self.edit_flight, underline=0)
        self.edit.add_command(label='Delete Flight', accelerator='Ctrl+D',
                              image=self.deleteimage, compound=tk.LEFT,
                              command=self.delete_flight, underline=0)
        self.top.add_cascade(label='Edit', menu=self.edit, underline=0)

        # create view menu
        # commands: flight details, statistics
        self.view = tk.Menu(self.top, tearoff=False)
        self.view.add_command(label='Flight Details', accelerator='Ctrl+F',
                              image=self.detailimage, compound=tk.LEFT,
                              command=self.switch_to_tab_1, underline=0)
        self.view.add_command(label='Statistics', accelerator=None,
                              image=self.statsimage, compound=tk.LEFT,
                              command=self.view_statistics, underline=0)
        self.view.add_separator()

        # create view/plot sub-menu
        # commands: flight route, airports
        self.plot = tk.Menu(self.view, tearoff=False)
        self.plot.add_cascade(label='Flight Route', accelerator=None,
                              image=self.worldimage, compound=tk.LEFT,
                              command=self.launch_routemap, underline=0)
        self.plot.add_cascade(label='Airports',
                              image=self.mapimage, compound=tk.LEFT,
                              command=self.launch_airport_map, underline=0)
        self.view.add_cascade(label='Plot', menu=self.plot)
        self.top.add_cascade(label='View', menu=self.view, underline=0)

        # create tools menu
        # commands: flight finder, clear fields
        self.tools = tk.Menu(self.top, tearoff=False)
        self.tools.add_command(label='Flight Finder', command=self.find_flight,
                               image=self.findimage, compound=tk.LEFT,
                               underline=0)
        self.tools.add_command(label='Clear Fields', accelerator='Ctrl+C',
                               image=self.clearimage, compound=tk.LEFT,
                               command=self.clear_fields, underline=0)
        self.tools.add_separator()

        # create tools/generate sub-menu
        # command: Spreadsheet, json file
        self.generate = tk.Menu(self.tools, tearoff=False)
        self.generate.add_command(label='Spreadsheet',
                                  accelerator='Ctrl+S',
                                  command=self.generate_spreadsheet,
                                  image=self.generateimage, compound=tk.LEFT,
                                  underline=0)
        self.generate.add_command(label='JSON file',
                                  accelerator='Ctrl+J',
                                  command=self.generate_json,
                                  image=self.jsonimage, compound=tk.LEFT,
                                  underline=0)
        self.tools.add_cascade(label='Generate', menu=self.generate)
        self.top.add_cascade(label='Tools', menu=self.tools, underline=0)

        # create help menu
        # commands: view help, about
        self.info = tk.Menu(self.top, tearoff=False)
        self.info.add_command(label='View Help', compound=tk.LEFT,
                              image=self.helpimage, command=self.view_help,
                              underline=0)
        self.info.add_command(label='About ...', command=self.view_credits,
                              compound=tk.LEFT, image=self.aboutimage,
                              underline=0)
        self.top.add_cascade(label='Help', menu=self.info, underline=0)

    def build_interface(self):
        """.Initializes and builds main program tab widgets."""
        # ------------------- create home notebook tab -------------------------

        # notebook frame
        self.frame_main = tk.Frame(self, background=None)
        self.frame_main.pack(side='top', expand='yes', fill='both', padx=0,
                             pady=0)
        # create notebook
        self.notebook = ttk.Notebook(self.frame_main)
        self.notebook.pack(side="top", expand="yes", fill="both", padx=0,
                           pady=0)
        # create main notebook tab
        self.tab_1 = tk.Frame(self.notebook)
        self.notebook.add(self.tab_1, text="Flight details ")

        # create key list and data field labelframe
        self.label_frame_main = tk.LabelFrame(self.tab_1)
        self.label_frame_main.pack(side='top', expand='yes', fill='both',
                                   padx=1, pady=1)

        # create key list descriptor box
        self.key_list_descriptor = tk.Label(self.label_frame_main, text="Keys",
                                            bd=1, relief="ridge", anchor="w")
        self.key_list_descriptor.grid(column=0, row=0, columnspan=2,
                                      sticky="we")

        # create database key listbox and related scrollbar
        self.key_list = tk.Listbox(self.label_frame_main, relief="sunken",
                                   height=8, width=22, bg='white')
        self.key_list.grid(column=0, row=1, rowspan=6, padx=0, pady=0)
        self.sbar = tk.Scrollbar(self.label_frame_main)
        self.key_list.config(yscrollcommand=self.sbar.set)
        self.sbar.config(command=self.key_list.yview)
        self.sbar.grid(column=1, row=1, rowspan=6, pady=0, sticky='wns')

        # create current flight data descriptor box
        self.key_list_descriptor = tk.Label(self.label_frame_main, text="Data",
                                            bd=1, relief="ridge", anchor="w")
        self.key_list_descriptor.grid(column=2, row=0, columnspan=2,
                                      sticky="we")

        # create current flight data labels
        self.lbl1 = tk.Label(self.label_frame_main, text='Flight Number: ')
        self.lbl1.grid(row=1, column=2, sticky="w")

        self.lbl2 = tk.Label(self.label_frame_main, text='Aircraft Model: ')
        self.lbl2.grid(row=2, column=2, sticky="w")

        self.lbl3 = tk.Label(self.label_frame_main, text='Flight Duration: ')
        self.lbl3.grid(row=3, column=2, sticky="w")

        self.lbl4 = tk.Label(self.label_frame_main, text='Air Carrier: ')
        self.lbl4.grid(row=4, column=2, sticky="w")

        self.lbl5 = tk.Label(self.label_frame_main, text='Departure: ')
        self.lbl5.grid(row=5, column=2, sticky="w")

        self.lbl6 = tk.Label(self.label_frame_main, text='Arrival: ')
        self.lbl6.grid(row=6, column=2, sticky="w")

        # create current flight data entry boxes
        self.entry1 = tk.Entry(self.label_frame_main, width=35,
                               state="disabled",
                               disabledforeground="black")
        self.entry1.grid(row=1, column=3)

        self.entry2 = tk.Entry(self.label_frame_main, width=35,
                               state="disabled",
                               disabledforeground="black")
        self.entry2.grid(row=2, column=3)

        self.entry3 = tk.Entry(self.label_frame_main, width=35,
                               state="disabled",
                               disabledforeground="black")
        self.entry3.grid(row=3, column=3)

        self.entry4 = tk.Entry(self.label_frame_main, width=35,
                               state="disabled",
                               disabledforeground="black")
        self.entry4.grid(row=4, column=3)

        self.entry5 = tk.Entry(self.label_frame_main, width=35,
                               state="disabled",
                               disabledforeground="black")
        self.entry5.grid(row=5, column=3)

        self.entry6 = tk.Entry(self.label_frame_main, width=35,
                               state="disabled",
                               disabledforeground="black")
        self.entry6.grid(row=6, column=3)

        # --------------------- create toolbar ---------------------------------
        self.label_frame_toolbar = tk.LabelFrame(self.frame_main)
        self.label_frame_toolbar.pack(side='top', expand='yes', fill='x',
                                      padx=1, pady=0)

        # create upload flight button
        self.new_flight_button = ttk.Button(self.label_frame_toolbar,
                                            image=self.uploadimage,
                                            command=self.upload_flight)
        self.new_flight_button.grid(row=0, column=0, padx=2)
        self.upload_tip = ToolTip(self.new_flight_button,
                                  ["Upload", "flight"])

        # create edit flight button
        self.edit_flight_button = ttk.Button(self.label_frame_toolbar,
                                             image=self.editimage,
                                             command=self.edit_flight)
        self.edit_flight_button.grid(row=0, column=1, padx=2)
        self.edit_tip = ToolTip(self.edit_flight_button,
                                ["Edit", "flight"])

        # create delete flight button
        self.delete_flight_button = ttk.Button(self.label_frame_toolbar,
                                               image=self.deleteimage,
                                               command=self.delete_flight)
        self.delete_flight_button.grid(row=0, column=2, padx=2)
        self.delete_tip = ToolTip(self.delete_flight_button,
                                  ["Delete", "flight", "(Ctrl+D)"])

        # create find flight button
        self.find_flight_button = ttk.Button(self.label_frame_toolbar,
                                             image=self.findimage,
                                             command=self.find_flight)
        self.find_flight_button.grid(row=0, column=3, padx=2)
        self.find_tip = ToolTip(self.find_flight_button,
                                ["Find", "flight"])

        # create clear fields button
        self.clear_button = ttk.Button(self.label_frame_toolbar,
                                       image=self.clearimage,
                                       command=self.clear_fields)
        self.clear_button.grid(row=0, column=4, padx=2)
        self.clear_tip = ToolTip(self.clear_button, ["Clear", "fields",
                                                     "(Ctrl+C)"])

        # create plot route button
        self.map_button = ttk.Button(self.label_frame_toolbar,
                                     image=self.worldimage,
                                     command=self.launch_routemap)
        self.map_button.grid(row=0, column=5, padx=2)
        self.about_tip = ToolTip(self.map_button, ["Plot", "route"])

        # -------------------- create status label -----------------------------
        self.frame_status = tk.Frame(self)
        self.frame_status.pack(anchor="e")

        font_status = ('verdana', 6, 'normal')
        self.label_status = tk.Label(self.frame_status, text="Idle...",
                                     font=font_status)
        self.label_status.pack()

        # update database key list, status label and fill iata list constant
        self.populate_list()
        self.get_last_access()
        self.build_iata_list()

        # Set keyboard bindings
        self.bind_all('<Control-c>', self.clear_fields)
        self.bind_all('<Control-C>', self.clear_fields)
        self.bind_all('<Control-d>', self.delete_flight)
        self.bind_all('<Control-D>', self.delete_flight)
        self.bind_all('<Control-e>', self.export_database)
        self.bind_all('<Control-E>', self.export_database)
        self.bind_all('<Control-f>', self.switch_to_tab_1)
        self.bind_all('<Control-F>', self.switch_to_tab_1)
        self.bind_all('<Control-j>', self.generate_json)
        self.bind_all('<Control-J>', self.generate_json)
        self.bind_all('<Control-i>', self.import_database)
        self.bind_all('<Control-I>', self.import_database)
        self.bind_all('<Control-s>', self.generate_spreadsheet)
        self.bind_all('<Control-S>', self.generate_spreadsheet)
        self.key_list.bind('<Return>', self.view_flight_data)
        self.key_list.bind('<Double-1>', self.view_flight_data)

    def populate_list(self):
        """Populates flight date listbox."""

        # update status label
        self.label_status['text'] = "Populating date list..."
        self.key_list.delete(0, tk.END)

        # extract list of keys from database
        database = sqlite3.connect('data/database.kr')
        try:
            cursor = database.cursor()
            dates = []
            pos = 0

            # fill dates list
            for date in cursor.execute('SELECT date FROM '
                                       'flight_data ORDER BY date'):
                dates.append(date[0])

            # insert dates list contents in key listbox
            for flight in dates:
                self.key_list.insert(pos, flight)
                pos += 1

            # color listbox rows in alternating shades
            for i in range(0, len(dates), 2):
                self.key_list.itemconfig(i, background='#f0f0ff')

        except sqlite3.Error:
            self.label_status['text'] = "Error..."
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
        finally:
            database.close()

    def activate_fields(self):
        """Switches all main notebook entry widgets statuses to tk.NORMAL"""
        self.entry1.config(state=tk.NORMAL)
        self.entry2.config(state=tk.NORMAL)
        self.entry3.config(state=tk.NORMAL)
        self.entry4.config(state=tk.NORMAL)
        self.entry5.config(state=tk.NORMAL)
        self.entry6.config(state=tk.NORMAL)

    def deactivate_fields(self):
        """Switches all main notebook entry widgets statuses to tk.Disabled"""
        self.entry1.config(state=tk.DISABLED)
        self.entry2.config(state=tk.DISABLED)
        self.entry3.config(state=tk.DISABLED)
        self.entry4.config(state=tk.DISABLED)
        self.entry5.config(state=tk.DISABLED)
        self.entry6.config(state=tk.DISABLED)

    def clear_fields(self, event=None):
        """Clears main tab text entries."""
        self.activate_fields()
        self.entry1.delete(0, tk.END)
        self.entry2.delete(0, tk.END)
        self.entry3.delete(0, tk.END)
        self.entry4.delete(0, tk.END)
        self.entry5.delete(0, tk.END)
        self.entry6.delete(0, tk.END)
        self.deactivate_fields()

        # update status
        self.label_status['text'] = "Idle..."

    def get_last_access(self):
        """ Loads last access Data from log.json file in data folder.

            If the file and Data folder don't exist the function creates them."""

        # check for log file existence
        if os.path.exists('data/log.json'):
            with open('data/log.json', 'r') as file:
                data = json.load(file)

            # update status label
            self.label_status['text'] = 'Welcome back! ' \
                                        'Your last access' \
                                        ' was on %s' % data['time']

            # free memory
            del data
        else:

            # display default status message
            self.label_status['text'] = 'Welcome to Kerosene! Your ' \
                                        'personal flight database'

    def view_flight_data(self, event=None):
        """ Returns flight data selected by the user in the listbox."""

        # update status label
        self.label_status['text'] = "Retrieving flight data ..."

        # connect to database
        database = sqlite3.connect('data/database.kr')
        cursor = database.cursor()
        try:

            # clear fields
            self.activate_fields()
            self.entry1.delete(0, tk.END)
            self.entry2.delete(0, tk.END)
            self.entry3.delete(0, tk.END)
            self.entry4.delete(0, tk.END)
            self.entry5.delete(0, tk.END)
            self.entry6.delete(0, tk.END)

            # get user selection from inbox
            index = self.key_list.curselection()
            selection = self.key_list.get(index)
            selection = str(selection)

            # fetch flight data
            fetch_row = cursor.execute('''SELECT * FROM flight_data
                                          WHERE date=?''', (selection,))
            record = fetch_row.fetchone()
            self.entry1.insert(0, record[1])  # flight number
            self.entry2.insert(0, record[2])  # plane model
            self.entry3.insert(0, record[5])  # flight duration
            self.entry4.insert(0, record[8])  # air carrier
            self.entry5.insert(0, record[6])  # departure
            self.entry6.insert(0, record[7])  # destination
            self.deactivate_fields()

            # update status label
            self.label_status['text'] = "Viewing flight data for {record}" \
                                        "...".format(record=record[1])

        # warn user if he doesn't select a key from the listbox
        except tk.TclError:
            pass
            self.label_status['text'] = "Error..."
            box.showwarning(title='No flight selected',
                            message='Select a flight from the keys list!')

        except sqlite3.Error:
            self.label_status['text'] = "Error..."
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
        finally:
            database.close()

    def upload_flight(self, event=None):
        """ Uploads new flight data to database.

            The function opens the UploaderEditorTab class in upload mode by
            setting is_uploader argument to True."""
        self.clear_fields()
        self.tab_upload = UploaderEditorTab(is_uploader=True)
        self.notebook.add(self.tab_upload, text="Upload flight  ")
        self.notebook.select(self.tab_upload)

        # disable menu options
        self.disable_menu_options()

        # focus on tab
        self.tab_upload.grab_set()
        self.tab_upload.focus()

    def edit_flight(self, event=None):
        """ Allows user to edit wrongly inputted flight data.

            The function opens the UploaderEditorTab class in editor mode by
            setting is_uploader argument to False."""
        self.clear_fields()
        try:

            # get key selected by the user in the key lostbox
            index = self.key_list.curselection()
            selection = self.key_list.get(index)
            selection = str(selection)

            # launch tab
            self.tab_edit = UploaderEditorTab(is_uploader=False,
                                              key=selection)
            self.notebook.add(self.tab_edit, text="Edit flight  ")
            self.notebook.select(self.tab_edit)

            # disable menu options
            self.disable_menu_options()

            # focus on tab
            self.tab_edit.grab_set()
            self.tab_edit.focus()

        # warn the user if he fails to open the tab without selecting a flight
        except tk.TclError:
            pass
            self.label_status['text'] = "Error..."
            box.showwarning(title='No flight selected',
                            message='Select a flight from the keys list!')
        finally:

            # update status label
            self.label_status['text'] = "Idle..."

    def delete_flight(self, event=None):
        """Deletes selected flight from database."""

        # update status label
        self.label_status['text'] = "Deleting flight data..."

        # set-up database
        database = sqlite3.connect('data/database.kr')
        cursor = database.cursor()
        try:

            # get user selection in listbox
            index = self.key_list.curselection()
            selection = self.key_list.get(index)
            selection = str(selection)

            # ask delete confirmation from user
            query = box.askyesno(title='Confirm',
                                 message='Are you sure you want to'
                                         ' delete this flight?')
            if query is True:

                # clear fields
                self.entry1.config(state=tk.NORMAL)
                self.entry2.config(state=tk.NORMAL)
                self.entry3.config(state=tk.NORMAL)
                self.entry4.config(state=tk.NORMAL)
                self.entry5.config(state=tk.NORMAL)
                self.entry6.config(state=tk.NORMAL)
                self.entry1.delete(0, tk.END)
                self.entry2.delete(0, tk.END)
                self.entry3.delete(0, tk.END)
                self.entry4.delete(0, tk.END)
                self.entry5.delete(0, tk.END)
                self.entry6.delete(0, tk.END)
                self.key_list.delete(0, tk.END)

                # delete flight record
                cursor.execute('''DELETE FROM flight_data
                        WHERE date=?''', (selection,))
                database.commit()

                # re-populate listbox with database keys
                self.populate_list()

                # disable main tab entries
                self.entry1.config(state=tk.DISABLED)
                self.entry2.config(state=tk.DISABLED)
                self.entry3.config(state=tk.DISABLED)
                self.entry4.config(state=tk.DISABLED)
                self.entry5.config(state=tk.DISABLED)
                self.entry6.config(state=tk.DISABLED)

                # alert the user that the flight is deleted
                box.showinfo(message='Flight deleted from database')
            else:
                pass

        # warn the user if he fails to select a key from the listbox
        except tk.TclError:
            pass
            self.label_status['text'] = "Error..."
            box.showwarning(title='No flight selected',
                            message='Select a flight from the keys list!')
        finally:
            database.close()
            self.populate_list()
            try:
                if kerosene.tab_stats.winfo_exists():

                    # update statistics if statistics tab exists
                    kerosene.tab_stats.update_data()
                else:
                    pass

            # ignore attribute error when statistics tab is not open
            except AttributeError:
                pass

            # update status label
            self.label_status['text'] = "Idle..."

    def view_statistics(self, event=None):
        """Launches statistics tab."""
        self.clear_fields()

        # launch statistics tab
        self.tab_stats = StatisticsTab()
        self.notebook.add(self.tab_stats, text="Statistics  ")
        self.notebook.select(self.tab_stats)

        # disable statistics menu options and grabs focus
        self.disable_menu_statistics()
        self.tab_stats.focus()

    def find_flight(self, event=None):
        """ Creates a small GUI to search for database keys in key listbox."""

        # set autocompletion list
        list_for_autocompletion = ['', ]
        for i, listbox_value in enumerate(self.key_list.get(0, tk.END)):
            list_for_autocompletion.append(listbox_value)

        # create toplevel window and adjust settings
        self.find_win = tk.Toplevel()
        self.find_win.resizable(width=False, height=False)
        self.find_win.title('Find')
        self.find_win.geometry('160x50+%d+%d' % (root.winfo_x() + 20,
                                                 root.winfo_y() + 20))
        self.find_win.iconbitmap('data/icons/magnifier.ico')

        # populate window with widgets
        self.label = tk.Label(self.find_win,
                              text='Find flight:').grid(row=0, column=0,
                                                        sticky='e')
        self.search_date = tk.StringVar()
        self.search_entry = AutocompleteEntry(self.find_win, width=15,
                                              textvariable=self.search_date)
        self.search_entry.set_completion_list(list_for_autocompletion)
        self.search_entry.grid(row=0, column=1, padx=0, pady=0, sticky='we')
        self.search_entry.focus_set()
        self.button = tk.Button(self.find_win, text='Search', underline=0,
                                command=self.search_for_flight)
        self.button.grid(row=1, column=1, sticky='se', padx=0, pady=0)

        # set focus on widget
        self.find_win.grab_set()
        self.find_win.focus()

        # set widget bindings
        self.search_entry.bind('<Return>', self.search_for_flight)

    def search_for_flight(self, event=None):
        """Flight date search engine.

           The function returns nothing if the flight date requested is not in
           the database."""

        # get user date
        query = self.search_date.get()

        # loop through listbox keys to find database key requested
        for i, listbox_entry in enumerate(self.key_list.get(0, tk.END), 0):
            if listbox_entry == query:

                # focus on selection and populate main tab entries with the
                # data requested, then clear search box
                self.key_list.selection_set(i)
                self.key_list.see(i)
                self.view_flight_data()
                self.search_entry.delete(0, tk.END)
            else:

                # delete search box text and ignore user query
                self.search_entry.delete(0, tk.END)

    def generate_routemap(self, event=None):
        """Generates a matplotlib map to visualise the selected flight's route."""

        # clear main tab entries and set status label
        self.clear_fields()
        self.label_status['text'] = "Building route map, the operation " \
                                    "may take a while..."
        try:

            # grab user selected key
            index = self.key_list.curselection()
            selection = self.key_list.get(index)
            selection = str(selection)

            # launch routemap tab
            self.tab_routemap = RoutemapTab(key=selection,
                                            is_pickled=self.map_pickle)
            self.notebook.add(self.tab_routemap, text="Routemap  ")
            self.notebook.select(self.tab_routemap)

            # focus on tab
            self.tab_routemap.grab_set()
            self.tab_routemap.focus()

            # disable menu options
            self.disable_menu_options()

        # warn the user if he has failed to select a key
        except tk.TclError:
            pass
            self.label_status['text'] = "Error..."
            box.showwarning(title='No flight selected',
                            message='Select a flight from the keys list!')

    def launch_routemap(self):
        """Threading method for generate_routemap function."""
        self.label_status['text'] = "Building route map..."
        run_thread = Thread(target=self.generate_routemap)
        run_thread.start()

    def generate_airport_map(self, event=None):
        """Generates a matplotlib map that plots all airports visited by user."""

        # clear main tab entries and set status label
        self.clear_fields()
        self.label_status['text'] = "Building airport map, the operation " \
                                    "may take a while..."

        # launch airport plot map tab
        self.tab_airportmap = RoutemapTab(is_route=False,
                                          is_pickled=self.map_pickle)
        self.notebook.add(self.tab_airportmap, text="Routemap  ")
        self.notebook.select(self.tab_airportmap)

        # focus on tab
        self.tab_airportmap.grab_set()
        self.tab_airportmap.focus()

        # disable menu options
        self.disable_menu_options()

    def launch_airport_map(self):
        """Threading method for launch_airport_map function."""
        self.label_status['text'] = "Building airport map..."
        run_thread = Thread(target=self.generate_airport_map)
        run_thread.start()

    def view_credits(self):
        """ Opens a new window providing credits information."""

        # launch window and configure window settings
        self.win_credits = CreditsTool(self)
        self.win_credits.resizable(width=False, height=False)
        self.win_credits.title('')
        self.win_credits.iconbitmap('data/icons/information.ico')
        self.win_credits.geometry('+%d+%d' % (root.winfo_x() +
                                              20, root.winfo_y() + 20))

        # set focus on window
        self.win_credits.grab_set()
        self.win_credits.focus()

        # start mainloop
        self.win_credits.mainloop()

    def import_database(self, event=None):
        """Imports data from other database.kr sqlite files.

           Allows users to conveniently import flight data using
           tkinter's filedialog facility."""

        # find path to desktop and open filedialog.askopenfilename
        user = os.path.expanduser('~')
        desktop = user + '\\Desktop'
        database_file = fd.askopenfilename(initialdir=desktop,
                                           defaultextension='.kr',
                                           filetypes=[('Kerosene datafile',
                                                       '*.kr')])

        # if a valid file is found alert the user about overwrite risks
        if database_file:
            query = box.askokcancel('Warning',
                                    'Importing a new database will erase'
                                    '\nprevious flight data, continue?')
            if query is True:

                # import selected database.kr file
                os.remove('data/database.kr')
                copy(database_file, 'data/')
                box.showinfo('Success', 'Database successfully imported')
                self.populate_list()
                try:
                    if kerosene.tab_stats.winfo_exists():

                        # update stats if tab exists
                        kerosene.tab_stats.update_data()
                    else:
                        pass

                # ignore attribute error when statistics tab is not open
                except AttributeError:
                    pass

    def switch_to_tab_1(self, event=None):
        """ Switches notebook focus to the program's main tab."""
        self.notebook.select(self.tab_1)

    def disable_menu_options(self):
        """Disables options on program menu.

           Disables options in File, Edit, View, Plot, Tools and Info submenus
           on program menu."""
        self.file.entryconfigure("Import Database", state="disabled")
        self.edit.entryconfigure("Upload Flight", state="disabled")
        self.edit.entryconfigure("Edit Flight", state="disabled")
        self.edit.entryconfigure("Delete Flight", state="disabled")
        self.view.entryconfigure("Flight Details", state="disabled")
        self.view.entryconfigure("Statistics", state="disabled")
        self.plot.entryconfigure("Flight Route", state="disabled")
        self.plot.entryconfigure("Airports", state="disabled")
        self.tools.entryconfigure("Flight Finder", state="disabled")
        self.info.entryconfigure("View Help", state="disabled")
        self.info.entryconfigure("About ...", state="disabled")

    def enable_menu_options(self):
        """Enables options on program menu.

           Re-enables options in File, Edit, View, Plot, Tools and Info submenus
           on program menu."""
        self.file.entryconfigure("Import Database", state="normal")
        self.edit.entryconfigure("Upload Flight", state="normal")
        self.edit.entryconfigure("Edit Flight", state="normal")
        self.edit.entryconfigure("Delete Flight", state="normal")
        self.view.entryconfigure("Flight Details", state="normal")
        self.view.entryconfigure("Statistics", state="normal")
        self.plot.entryconfigure("Flight Route", state="normal")
        self.plot.entryconfigure("Airports", state="normal")
        self.tools.entryconfigure("Flight Finder", state="normal")
        self.info.entryconfigure("View Help", state="normal")
        self.info.entryconfigure("About ...", state="normal")

    def disable_menu_statistics(self):
        """Disables view statistics option on program menu."""
        self.view.entryconfigure("Statistics", state="disabled")

    def enable_menu_statistics(self):
        """Enables view statistics option on program menu."""
        self.view.entryconfigure("Statistics", state="normal")

    @staticmethod
    def pickle_map():
        """Generates the base matplotlib map used by the program and caches it."""
        matplotlib.use('TkAgg')
        fig = Figure(figsize=(6, 3))

        # set map background color
        fig.patch.set_facecolor('lightgray')

        ax1 = fig.add_subplot(111)
        m = Basemap(projection='robin',
                    lat_0=43.0000, lon_0=12.0000,
                    resolution='c', area_thresh=100000,
                    ax=ax1)

        # build map and adjust settings
        m.drawcoastlines(linewidth=1, color="black")
        m.drawcountries(color="#ddaa66")
        m.fillcontinents(color='#ddaa66', lake_color="#b0c4de")
        m.drawmapboundary(linewidth=2, color="black", fill_color="#b0c4de")
        m.drawmeridians(np.arange(0, 360, 30), color="black")
        m.drawparallels(np.arange(-90, 90, 30), color="black")
        # m.bluemarble()

        return pickle.dumps((m, fig))

    @staticmethod
    def view_help():
        """Launches program help file."""
        os.startfile("docs\\kerosene_docs.chm")
        # TODO: help file needs to be updated to take into account added functionalities; like plot maps for instance

    @staticmethod
    def string_to_timedelta(string=''):
        """This function converts datetime string representations saved in
        the database."""
        split_string = string.split(':')
        converted_string = dt.timedelta(hours=int(split_string[0]),
                                        minutes=int(split_string[1]),
                                        seconds=int(split_string[2]))
        return converted_string

    def generate_spreadsheet(self, event=None):
        """Allows users to export all flight data to .xlsx spreadsheet
        format, the file is saved on the desktop."""

        # update status label and set database
        self.label_status['text'] = "Generating spreadsheet..."
        database = sqlite3.connect('data/database.kr')
        try:
            cursor = database.cursor()

            # find desktop path
            user = os.path.expanduser('~')
            desktop = user + '\\Desktop'
            wb = Workbook()
            ws1 = wb.active

            # set worbook row 1 cell titles
            ws1.cell('A1').value = 'Date'
            ws1.cell('B1').value = 'Number'
            ws1.cell('C1').value = 'Model'
            ws1.cell('D1').value = 'Duration'
            ws1.cell('E1').value = 'Carrier'
            ws1.cell('F1').value = 'Departure'
            ws1.cell('G1').value = 'Destination'
            ws1.cell('H1').value = 'Departure IATA'
            ws1.cell('I1').value = 'Destination IATA'
            ws1.cell('J1').value = 'Departure latitude'
            ws1.cell('K1').value = 'Departure longitude'
            ws1.cell('L1').value = 'Departure city'
            ws1.cell('M1').value = 'Destination city'
            ws1.cell('N1').value = 'Destination latitude'
            ws1.cell('O1').value = 'Destination longitude'

            # loop through each record and assign variables to data
            for i, flight in enumerate(cursor.execute('SELECT * FROM '
                                                      'flight_data')):
                date = flight[0]
                flight_number = flight[1]
                plane_model = flight[2]
                flight_duration = flight[5]
                carrier = flight[8]
                departure = flight[6]
                destination = flight[7]
                iata_dep = flight[9]
                iata_des = flight[10]
                latitude_dep = flight[11]
                longitude_dep = flight[12]
                city_dep = flight[13]
                city_des = flight[14]
                latitude_des = flight[15]
                longitude_des = flight[16]

                # paste data in cells
                ws1.cell(row=i + 2, column=1).value = date
                ws1.cell(row=i + 2, column=2).value = flight_number
                ws1.cell(row=i + 2, column=3).value = plane_model
                ws1.cell(row=i + 2, column=4).value = flight_duration
                ws1.cell(row=i + 2, column=5).value = carrier
                ws1.cell(row=i + 2, column=6).value = departure
                ws1.cell(row=i + 2, column=7).value = destination
                ws1.cell(row=i + 2, column=8).value = iata_dep
                ws1.cell(row=i + 2, column=9).value = iata_des
                ws1.cell(row=i + 2, column=10).value = latitude_dep
                ws1.cell(row=i + 2, column=11).value = longitude_dep
                ws1.cell(row=i + 2, column=12).value = city_dep
                ws1.cell(row=i + 2, column=13).value = city_des
                ws1.cell(row=i + 2, column=14).value = latitude_des
                ws1.cell(row=i + 2, column=15).value = longitude_des

            # save file on desktop
            wb.save(filename=desktop + '\\kerosene_data.xlsx')

            # alert user to successfull data transfer
            box.showinfo('Success!', "Spreadsheet generated. "
                                     "\nThe file has been saved on your Desktop.")
        except sqlite3.Error as e:
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
        finally:

            # close database and update status label
            database.close()
            self.label_status['text'] = "Idle..."

    def generate_json(self, event=None):
        """Allows users to export all flight data to .json format, the file is
        saved on the desktop."""

        # update status label, connect to database and set database dictionary
        self.label_status['text'] = "Exporting data to JSON..."
        database = sqlite3.connect('data/database.kr')
        json_database = {}
        try:
            cursor = database.cursor()

            # find desktop path
            user = os.path.expanduser('~')
            desktop = user + '\\Desktop'

            # loop through each record and assign variables to data
            for flight in cursor.execute('SELECT * FROM flight_data'):
                date = flight[0]
                flight_number = flight[1]
                plane_model = flight[2]
                take_off_time = flight[3]
                landing_time = flight[4]
                flight_duration = flight[5]
                carrier = flight[8]
                departure = flight[6]
                destination = flight[7]
                iata_dep = flight[9]
                iata_des = flight[10]
                latitude_dep = flight[11]
                longitude_dep = flight[12]
                city_dep = flight[13]
                city_des = flight[14]
                latitude_des = flight[15]
                longitude_des = flight[16]

                # build a dictionary copy of the current flight record
                record = {'date': date,
                          'flight_number': flight_number,
                          'plane_model': plane_model,
                          'take_off_time': take_off_time,
                          'landing_time': landing_time,
                          'flight_duration': flight_duration,
                          'departure': departure,
                          'destination': destination,
                          'carrier': carrier,
                          'iata_departure': iata_dep,
                          'iata_destination': iata_des,
                          'latitude_departure': latitude_dep,
                          'longitude_departure': longitude_dep,
                          'city_departure': city_dep,
                          'city_destination': city_des,
                          'latitude_destination': latitude_des,
                          'longitude_destination': longitude_des}

                # upload dictionary to master json dictionary
                json_database[date] = record

            # save json file to the user's desktop
            with open(desktop + '/kerosene_data.json', 'w') as file:
                json.dump(json_database, file)

            # alert user that the data copy was successfull
            box.showinfo('Success!', "JSON file generated. "
                                     "\nThe file has been saved on your Desktop.")
        except sqlite3.Error as e:
            box.showwarning('Error', 'Oops, something went wrong!'
                                     '\nThe database appears not'
                                     '\nto be working properly.')
        finally:

            # close database and update status label
            database.close()
            self.label_status['text'] = "Idle..."

    @staticmethod
    def export_database(event=None):
        """ Function allowing users to export/backup their database.kr file.

            The database file is stored on the desktop in a folder named
            'kerosene_backup'."""

        # find user desktop path
        user = os.path.expanduser('~')
        desktop = user + '\\Desktop'

        # see if another backup copy exists
        if os.path.exists(desktop + "\\kerosene_backup"):
            box.showwarning('Error', 'Existing backup detected!')

        # create backup copy and save to desktop
        else:
            os.mkdir(desktop + '\\kerosene_backup')
            copy('data/database.kr', desktop + '\\kerosene_backup')
            box.showinfo('Success', 'Backup created!')

    @staticmethod
    def format_timedelta(time_to_format):
        """Formats timedelta objects.

           Method displays timedelta time in a HOURS:MINUTES:SECONDS format."""
        minutes, seconds = divmod(time_to_format.seconds +
                                  time_to_format.days * 86400, 60)
        hours, minutes = divmod(minutes, 60)
        return '{:d}:{:02d}:{:02d}'.format(hours, minutes, seconds)

    @staticmethod
    def quit_program():
        """Method to safely quit the program.

           The method records time and date in log.json file for use on
           status label the next time the program is launched """

        # open log.json file
        with open('data/log.json', 'w') as file:
            # extract and format current time
            now = dt.datetime.now()
            current_time = now.strftime("%a, %d %b %Y at %H:%M")

            # save current date and time
            data = {'time': current_time}
            json.dump(data, file)

        # quit program
        root.destroy()

    @staticmethod
    def build_iata_list():
        """Threading method for upload_iatas function."""

        # invoke global constant
        global IATA_CODE_LIST

        # connect to database
        database = sqlite3.connect('data/airports_data.sqlite')
        cursor = database.cursor()
        iatas = cursor.execute('SELECT iata FROM airports')
        sorted_iatas = (iata[0] for iata in sorted(iatas))

        # loop through IATA codes and append to global iata code list
        for iata in sorted_iatas:

            # ignore first '---' iata field
            if iata == '---':
                pass
            else:
                IATA_CODE_LIST.append(iata)

        # close database
        database.close()


# Launch program ---------------------------------------------------------------
if __name__ == "__main__":
    root = tk.Tk()
    root.resizable(width=False, height=False)
    root.title("Kerosene 1.0.0")
    kerosene = Gui(root)
    root.config(menu=kerosene.top)
    kerosene.pack(expand='yes', fill='both')
    root.iconbitmap('data/icons/flame.ico')
    root.protocol('WM_DELETE_WINDOW', Gui.quit_program)
    root.mainloop()
